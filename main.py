# -*- coding: utf8 -*-
import sys
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QApplication, QMainWindow
import json
import os
from design import Ui_Main
import openpyxl
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, NamedStyle
from openpyxl.styles.colors import Color


def file_excel():
    with open ('data.json','r') as file:
        data = json.load(file)
    file_name = data['pasport_ishodnie_dannye']['file_name']
    file_path = data['pasport_ishodnie_dannye']['path_for_document']
    os.chdir(file_path)



class Main(QMainWindow, Ui_Main):
    def __init__(self, parent=None):
        super(Main, self).__init__(parent)
        self.setupUi(self)
        
        self.pushButton1.clicked.connect(self.ishodnie_dannye)
        self.pushButton2.clicked.connect(self.pasport_AFS_1)
        self.pushButton3.clicked.connect(self.jurnal_AFS)
        


    def main_menu(self):
        self.QtStack.setCurrentIndex(0)
        



########                              ############
########                              ############
#       страница с исходными данными
########                              ############
########                              ############

    def ishodnie_dannye(self):
        self.QtStack.setCurrentIndex(1)
        
        def getDirectory_logs():
            dirlist = QtWidgets.QFileDialog.getExistingDirectory()
            self.plainTextEdit_folder.setPlainText(format(dirlist))

        def getDirectory_path_document():
            dirlist = QtWidgets.QFileDialog.getExistingDirectory()
            self.plainTextEdit_folder_4.setPlainText(format(dirlist))

        def btn_click():
            # СЧИТЫВАНИЕ ДАННЫХ
            
            folder_with_logs = self.plainTextEdit_folder.toPlainText() # папка с логами
            operator = self.plainTextEdit_folder_2.toPlainText() # ФИО оператора 
            object_name = self.plainTextEdit_folder_3.toPlainText() # наименование объекта
            path_for_document = self.plainTextEdit_folder_4.toPlainText() # путь до создоваемого файла
            file_name = self.plainTextEdit_folder_5.toPlainText() # название excel файла
            
            pasport = {'pasport_ishodnie_dannye' : {
            "folder_with_logs" : folder_with_logs,
            "operator" : operator,
            "object_name" : object_name,
            "path_for_document" : path_for_document,
            "file_name" : file_name,
            }}
            with open("data.json","w") as write_file: 
                json.dump(pasport,write_file) 
            self.main_menu()
            


        if os.path.exists('data.json')==True:
            try:   
                with open('data.json') as file:
                    file_content = file.read().strip()
            # Проверяем, пустой ли файл
                if not file_content:
                    pasport = {'pasport_ishodnie_dannye' : {
                        "folder_with_logs" : '',
                        "operator" : '',
                        "object_name" : '',
                        "path_for_document" : '',
                        "file_name" : '',
                        }}
                    with open("data.json","w") as write_file: 
                        json.dump(pasport,write_file) 
            
            
            except FileNotFoundError:
                pass


            
        else:
            pasport = {'pasport_ishodnie_dannye' : {
                        "folder_with_logs" : '',
                        "operator" : '',
                        "object_name" : '',
                        "path_for_document" : '',
                        "file_name" : '',
                        }}
            with open("data.json","w") as write_file: 
                json.dump(pasport,write_file) 

        try:
            with open("data.json", "r") as read_file:
                data = json.load(read_file)
        except:
            pass


        try:
            self.plainTextEdit_folder.setPlainText(data["pasport_ishodnie_dannye"]["folder_with_logs"])
            self.plainTextEdit_folder_2.setPlainText(data["pasport_ishodnie_dannye"]["operator"])
            self.plainTextEdit_folder_3.setPlainText(data["pasport_ishodnie_dannye"]["object_name"])
            self.plainTextEdit_folder_4.setPlainText(data["pasport_ishodnie_dannye"]["path_for_document"])
            self.plainTextEdit_folder_5.setPlainText(data["pasport_ishodnie_dannye"]["file_name"])
        except:
            self.plainTextEdit_folder.setPlainText("")
            self.plainTextEdit_folder_2.setPlainText("")
            self.plainTextEdit_folder_3.setPlainText("")
            self.plainTextEdit_folder_4.setPlainText("")
            self.plainTextEdit_folder_5.setPlainText("")

        self.pushButton_folder.clicked.connect(getDirectory_logs)
        self.pushButton_folder_2.clicked.connect(getDirectory_path_document)
        self.pushButton_download.clicked.connect(btn_click)



########                              ############
########                              ############
#         страница с паспортом 1
########                              ############
########                              ############

#сделать проверку есть ли уже такой афс или нет
    def pasport_AFS_1(self):
        self.QtStack.setCurrentIndex(2)
        self.pushButton.clicked.connect(self.main_menu)

        def btn_next():
            global Mission_number
            Mission_number = self.comboBox.currentText()
            Date = self.plainTextEdit_data.toPlainText()
            Time = self.plainTextEdit_time.toPlainText()
            AFS_type = self.comboBox_2.currentText()
            AFS_mode = self.comboBox_3.currentText()
            UMA_name = self.comboBox_4.currentText()
            page_one = {'AFS_' + Mission_number:{
                "Mission_number":Mission_number,
                "Date":Date,
                "Time":Time,
                "AFS_type":AFS_type,
                "AFS_mode":AFS_mode,
                "UMA_name":UMA_name,
            }}

            with open ('data.json','r') as file:
                data = json.load(file)
            data.update(page_one)
            with open('data.json',"w") as file:
                json.dump(data,file)
            self.pasport_AFS_2()


        self.pushButton_2.clicked.connect(btn_next)


########                              ############
########                              ############
#         страница с паспортом 2
########                              ############
########                              ############


    def pasport_AFS_2(self):
        self.QtStack.setCurrentIndex(3)
        self.pushButton_AFS_2.clicked.connect(self.pasport_AFS_1)

 
        def btn_next():
            registry_number = self.plainTextEdit_AFS_2.toPlainText()
            pay_load_1 = self.plainTextEdit_AFS_3.toPlainText()
            pay_load_2 = self.plainTextEdit_AFS_4.toPlainText()
            mission_software = self.plainTextEdit_AFS_5.toPlainText()
            solution_method = self.comboBox_AFS_2_1.currentText()
            altitude = self.plainTextEdit_AFS_6.toPlainText()
            page_two = {
                "registry_number":registry_number,
                "pay_load_1":pay_load_1,
                "pay_load_2":pay_load_2,
                "mission_software":mission_software,
                "solution_method":solution_method,
                "altitude":altitude,
            }
  
            with open ('data.json','r') as file:
                data = json.load(file)
            data['AFS_' + Mission_number].update(page_two)
            with open('data.json',"w") as file:
                json.dump(data,file)
            self.pasport_AFS_3()

        self.pushButton_AFS_2_2.clicked.connect(btn_next)
        


########                              ############
########                              ############
#         страница с паспортом 3
########                              ############
########                              ############


    def pasport_AFS_3(self):
        self.QtStack.setCurrentIndex(4)
        self.pushButton_AFS_3_1.clicked.connect(self.pasport_AFS_2)


        def btn_next():
            horizontal_lap  = self.plainTextEdit_AFS_3_1.toPlainText()
            vertical_lap  = self.plainTextEdit_AFS_3_3.toPlainText()
            shape = self.plainTextEdit_AFS_3_2.toPlainText()
            shots_number = self.plainTextEdit_AFS_3_4.toPlainText()
            precipitation = self.comboBox_AFS_3_1.currentText()
            undercast = self.comboBox_AFS_3_2.currentText()
            page_three = {
                "horizontal_lap":horizontal_lap,
                "vertical_lap":vertical_lap,
                "shape":shape,
                "shots_number":shots_number,
                "precipitation":precipitation,
                "undercast":undercast,
            }
  
            with open ('data.json','r') as file:
                data = json.load(file)
            data['AFS_' + Mission_number].update(page_three)
            with open('data.json',"w") as file:
                json.dump(data,file)
            self.geodeziy()


        self.pushButton_AFS_3_2.clicked.connect(btn_next)


########                              ############
########                              ############
#         страница с геодезией
########                              ############
########                              ############

    def geodeziy(self):
        self.QtStack.setCurrentIndex(5)
        self.pushButton_geodeziy_1.clicked.connect(self.pasport_AFS_3)

        def btn_next():
            home_point  = self.plainTextEdit_geodeziy_1.toPlainText()
            device   = self.plainTextEdit_geodeziy_3.toPlainText()
            log_number = self.plainTextEdit_geodeziy_2.toPlainText()
            device_high = self.plainTextEdit_geodeziy_4.toPlainText()
            file_name  = self.plainTextEdit_geodeziy_5.toPlainText()
            
            page_four = {
                "home_point":home_point,
                "device":device,
                "log_number":log_number,
                "device_high":device_high,
                "file_name":file_name,
            }
  
            with open ('data.json','r') as file:
                data = json.load(file)
            data['AFS_' + Mission_number].update(page_four)
            with open('data.json',"w") as file:
                json.dump(data,file)
            self.primechania()

        self.pushButton_geodeziy_2.clicked.connect(btn_next)

########                              ############
########                              ############
#         страница с примечаниями
########                              ############
########                              ############

    def primechania(self):
        self.QtStack.setCurrentIndex(6)
        self.pushButton_primechania_1.clicked.connect(self.geodeziy)

        def btn_next():
            processing_usage  = self.comboBox_primechania.currentText()
            usage_problem   = self.plainTextEdit_primechania_2.toPlainText()
            incidents = self.plainTextEdit_primechania_1.toPlainText()

            
            page_five = {
                "processing_usage":processing_usage,
                "usage_problem":usage_problem,
                "incidents":incidents,
            }
  
            with open ('data.json','r') as file:
                data = json.load(file)
            data['AFS_' + Mission_number].update(page_five)
            with open('data.json',"w") as file:
                json.dump(data,file)
            create_excel_doc()
            add_ws()
            self.main_menu()
            
        # создание таблицы excel
        def create_excel_doc():
            file_name = ''
            file_path = ''
            with open ('data.json','r') as file:
                data = json.load(file)
                file_name = data["pasport_ishodnie_dannye"]["file_name"]
                file_path = data["pasport_ishodnie_dannye"]["path_for_document"]
            file_name = file_name + ".xlsx"
            if os.path.exists(file_path+"/"+file_name):
                # файл существует
                #print("Файл существует")
                pass
            else:
                # файл не существует
                os.chdir(file_path)
                workbook = xlsxwriter.Workbook(file_name)
                workbook.close()
                
                
        # добавление листа в таблицу (отдельный АФС)
        def add_ws():
            file_name = ''
            file_path = ''
            with open ('data.json','r') as file:
                data = json.load(file)
                file_name = data["pasport_ishodnie_dannye"]["file_name"]
                file_path = data["pasport_ishodnie_dannye"]["path_for_document"]
            file_name = file_name + ".xlsx"
            os.chdir(file_path)
            try:
                myfile = open(file_name, "r+")
            except IOError:
                print("Файл открыт")
            else:
                workbook = load_workbook(file_name)
                worksheet_AFS1 = workbook.create_sheet("АФС_1")
                row = 1
                column = 1
                frame_1 = ["Наименование объекта", "Оператор", "Номер полетного задания", "Дата полета", "Время полета", "Тип АФС", "Вид БВС", "Название БВС", "Регистрационный номер борта", "Полезная нагрузка 1", "Полезная нагрузка 2", "Метод решения", "Высота полета", "Продольное перекрытие", "Поперечное перекрытие", "Разрешение", "Количество снимков", "Осадки", "Облачность"]
                frame_2 = ["Геодезия", "Наименование точки (базы)", "Прибор (название, номер)", "Порядковый номер лога (базы)", "Высота прибора", "Название файла"]
                frame_3 = ["Примечания", "Использование полета в обработке", "Причина, по которой нельзя использовать", "Происшествия"]
                
                for item in frame_1 :
                    worksheet_AFS1.cell(row=row, column=column).value = item
                    row += 1
                
                row = 22
                column = 1
                for item in frame_2 :
                    worksheet_AFS1.cell(row=row, column=column).value = item
                    row += 1
                
                row = 22
                column = 4
                for item in frame_3 :
                    worksheet_AFS1.cell(row=row, column=column).value = item
                    row += 1
                
                worksheet_AFS1.merge_cells('D25:D27')
                worksheet_AFS1.merge_cells('E25:E27')
                worksheet_AFS1.merge_cells('C22:C27')
                worksheet_AFS1.merge_cells('A22:B22')
                worksheet_AFS1.merge_cells('D22:E22')
                
                highlight_blue = NamedStyle(name="highlight_blue")
                highlight_blue.font = Font(name = 'Arial', bold=True, size=11)
                blueFill = PatternFill(start_color='0099CCFF',
                   end_color='0099CCFF',
                   fill_type='solid')
                highlight_blue.fill = blueFill
                workbook.add_named_style(highlight_blue)
                
                usual_style = NamedStyle(name="usual_style")
                usual_style.font = Font(name = 'Arial', bold=False, size=11)
                usual_style.alignment = Alignment(horizontal='left',vertical='center', wrap_text=False, shrink_to_fit=False)
                workbook.add_named_style(usual_style)
                
                for c in range(1, 6) :
                    for r in range(1, 28):
                        worksheet_AFS1.cell(row=r, column=c).style = usual_style
                
                
                worksheet_AFS1['A22'].style = 'highlight_blue'
                worksheet_AFS1['D22'].style = 'highlight_blue'
                
                
                
                
                workbook.save(filename=file_name)
                workbook.close()

        self.pushButton_primechania_2.clicked.connect(btn_next)


########                              ############
########                              ############
#         страница с геодезией
########                              ############
########                              ############

    def jurnal_AFS(self):
        self.QtStack.setCurrentIndex(7)
        self.pushButton_jurnal.clicked.connect(self.main_menu)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    showMain = Main()
    sys.exit(app.exec_())
